# Importing libraries 
 
import pandas as pd
import numpy as np
import os
from openpyxl.workbook import Workbook   #used to save to our Excel sheet
from openpyxl import load_workbook

fileName = '2022-0000 Sample PM Tool.xlsx'
filePath = r'C:\Users\mary\Downloads'
file=os.path.join(filePath,fileName)
df=pd.read_excel(file)
df
to_drop = ['Unnamed: 1', 'Unnamed: 2', 'Unnamed: 3','Unnamed: 4', 'Unnamed: 5', 'Unnamed: 6', 'Unnamed: 8', 'Unnamed: 9', 'Unnamed: 10', 'Unnamed: 11', 'Unnamed: 13', 'Unnamed: 14', 'Unnamed: 16', 'Unnamed: 17', 'Unnamed: 19', 'Unnamed: 20', 'Unnamed: 21', 'Unnamed: 22', 'Unnamed: 23', 'Unnamed: 24', 'Unnamed: 28', 'Unnamed: 30', 'Unnamed: 32', 'Unnamed: 33', 'Unnamed: 34', 'Unnamed: 35', 'Unnamed: 36', 'Unnamed: 37', 'Unnamed: 39', 'Unnamed: 40','Unnamed: 41']
df.drop(columns=to_drop, inplace=True)
df.drop([0,1,2,3,4,5,10,11,12,13,14,15,16,17,18])
print(df)

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('January', 0)
ws2 = wb.create_sheet('February', 1)
ws3 = wb.create_sheet('March', 2)
ws4 = wb.create_sheet('April', 3)
ws5 = wb.create_sheet('May', 4)
ws6 = wb.create_sheet('June', 5)
ws7 = wb.create_sheet('July', 6)
ws8 = wb.create_sheet('August',7 )
ws9 = wb.create_sheet('September', 8)
ws10 = wb.create_sheet('October', 9)
ws11 = wb.create_sheet('November' 10)
ws12 = wb.create_sheet('December' 11)

wb2 = load_workbook('2022-0000 Sample PM Tool.xlsx')
new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active
