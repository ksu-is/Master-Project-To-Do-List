# Importing libraries
import numpy as np    
import pandas as pd
import xlrd
import xdrlib, sys
import os
import openpyxl
from openpyxl import load_workbook

# Raw data path
file_path = './data'
# List for excel title names
excel_title = []

# Selected row and column
startcol = 8
startrow = 9

# Read excel title names from local files
def getFileNames(file_path):
    """
    get local excel file names;
    extract file number from file name and append to list
    in order, and append to excel_title list
    """
    filenames = os.listdir(file_path)
    for i, filename in enumerate(filenames):
        if i == 0:
            iSpecialFile = i+1
            sFileName = filename
        excel_title.append(filename)
    return excel_title
getFileNames(file_path)

# Append dataframe to excel sheets
def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, **to_excel_kwargs):
    """
    paste dataframe to excel sheet
    at selected region (col=.., row=..);
    """
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    try:
        writer.book = load_workbook(filename)
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        pass
    if startrow is None:
        startrow = 0
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
    writer.save()

# Read some sheet in each excel file
# Append dataframe to excel_name list
excel_name = []
for i in range(len(excel_title)):
    df = pd.read_excel( file_path+'/'+excel_title[i], 'sheet1')
    excel_name.append(df)

# Column name in sheet1
col_name = ['column1', 'column2', 'column3', 'column4', 'column5',\
            'column6', 'column7', 'column8', 'column9', 'column10']
    
# Combine the same dimension data in different input files
excel_num = len(excel_name)
col_num = len(col_name)  
value = []
for i in range(col_num):
    for j in range(excel_num):
        templist = excel_name[j][col_name[i]][1:].values.tolist()
        value.append(templist)

# Split excel title name
# e.g., split 'AA-BB-12345678_9876543210' and get '12345678'
k = []
for s in excel_title:
    temp1 = str(s).split('-')
    k.append(temp1[2].split('_')[0])  

# Output excel file and sheet name list
Output_file = "output.xlsx"
lst_sheet_name = ['Sheet 1', 'Sheet 2', 'Sheet 3', 'Sheet 4', 'Sheet 5',\
                  'Sheet 6', 'Sheet 7', 'Sheet 8']

# Paste the dataframe in excel_name (each combined dimension data)\
# to the corresponding Sheet in output file 
v = [] 
start = 0
for i in range(col_num):
    v = value[start:start+excel_num]
    start += excel_num
    dict_data = dict(zip(k, v))
    df_temp = pd.DataFrame.from_dict(dict_data, orient='index')
    df_temp1 = df_temp.transpose()
    #df_temp1.to_excel('XXX'+str(str(col_name[i]).split('_')[2])+'.xlsx',\
    #             encoding='utf-8-sig', index=False)
    append_df_to_excel(Output_file, df_temp1, sheet_name=lst_sheet_name[i],\
                       startcol=startcol, startrow=startrow, index=False)
    
